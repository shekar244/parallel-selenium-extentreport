import os
import pandas as pd
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.chart import (
    PieChart, BarChart, LineChart, Reference
)
from openpyxl.utils.dataframe import dataframe_to_rows

REPROCESS_ALL = True
DATETIME_FORMAT = "%m/%d/%Y %H:%M:%S"

def reset_processed_log(log_path):
    if os.path.exists(log_path):
        os.remove(log_path)

def load_processed_files(log_path):
    if os.path.exists(log_path):
        return set(pd.read_csv(log_path)['filename'])
    return set()

def save_processed_file(log_path, filename):
    mode = 'a' if os.path.exists(log_path) else 'w'
    header = not os.path.exists(log_path)
    with open(log_path, mode) as f:
        if header:
            f.write("filename\n")
        f.write(f"{filename}\n")

def extract_tests_from_html(html_path):
    with open(html_path, "r", encoding="utf-8", errors="ignore") as f:
        soup = BeautifulSoup(f, "html.parser")

    test_data = []
    for test in soup.select("ul.test-collection > li"):
        test_name = test.select_one("span.test-name")
        test_status = test.select_one("span.test-status")
        test_start = test.select_one("span.test-started-time")
        test_end = test.select_one("span.test-ended-time")

        name = test_name.text.strip() if test_name else "Unknown"
        status = test_status.text.strip() if test_status else "Unknown"
        start = test_start.text.strip() if test_start else ""
        end = test_end.text.strip() if test_end else ""

        if not end.strip():
            timestamps = test.select("td.timestamp")
            if timestamps and start:
                test_date = start.split()[0]
                last_time = timestamps[-1].text.strip()
                end = f"{test_date} {last_time}"

        test_data.append({
            "HTML Report": os.path.basename(html_path),
            "Test Name": name,
            "Test Status": status,
            "Start Time": start,
            "End Time": end
        })
    return test_data

def summarize_test_durations(df):
    df["Start Time"] = pd.to_datetime(df["Start Time"], errors="coerce")
    df["End Time"] = pd.to_datetime(df["End Time"], errors="coerce")
    df = df.dropna(subset=["Start Time", "End Time"])
    df["Duration (mins)"] = ((df["End Time"] - df["Start Time"]).dt.total_seconds() / 60).clip(lower=0).round(2)
    df["Start+Duration"] = list(zip(df["Start Time"], df["Duration (mins)"]))

    grouped = df.groupby("Test Name")["Start+Duration"].agg(
        SortedRuns=lambda x: [d for _, d in sorted(x)]
    ).reset_index()

    agg_stats = df.groupby("Test Name")["Duration (mins)"].agg(
        Count="count", Total_Time="sum", Max_Time="max"
    ).reset_index()

    merged = pd.merge(grouped, agg_stats, on="Test Name")
    max_runs = merged["SortedRuns"].apply(len).max()
    run_columns = pd.DataFrame(merged["SortedRuns"].tolist(),
                               columns=[f"run-{i+1}" for i in range(max_runs)])

    summary_df = pd.concat([
        merged[["Test Name", "Count", "Total_Time", "Max_Time"]],
        run_columns
    ], axis=1)

    return summary_df

def add_charts_to_workbook(excel_path, df_details, df_summary):
    wb = load_workbook(excel_path)
    ws = wb.create_sheet("charts")

    status_data = df_details["Test Status"].value_counts().reset_index()
    status_data.columns = ["Test Status", "Count"]
    for row in dataframe_to_rows(status_data, index=False, header=True):
        ws.append(row)

    pie = PieChart()
    pie.title = "Test Status Distribution"
    pie.add_data(Reference(ws, min_col=2, min_row=2, max_row=len(status_data)+1))
    pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(status_data)+1))
    ws.add_chart(pie, "E2")

    start_row = len(status_data) + 4
    ws.append([])
    ws.append(["Test Name", "Total Duration"])
    for row in df_summary[["Test Name", "Total_Time"]].itertuples(index=False):
        ws.append(list(row))

    bar = BarChart()
    bar.title = "Total Duration by Test Name"
    bar.x_axis.title = "Test Name"
    bar.y_axis.title = "Total Duration (mins)"
    bar.add_data(Reference(ws, min_col=2, min_row=start_row+2, max_row=start_row+1+len(df_summary)))
    bar.set_categories(Reference(ws, min_col=1, min_row=start_row+2, max_row=start_row+1+len(df_summary)))
    ws.add_chart(bar, f"E{start_row+2}")

    col_row_start = start_row + len(df_summary) + 6
    ws.append([])
    ws.append(["Test Name", "Run Count"])
    for row in df_summary[["Test Name", "Count"]].itertuples(index=False):
        ws.append(list(row))

    col_chart = BarChart()
    col_chart.title = "Test Run Count"
    col_chart.x_axis.title = "Test Name"
    col_chart.y_axis.title = "Run Count"
    col_chart.add_data(Reference(ws, min_col=2, min_row=col_row_start+2, max_row=col_row_start+1+len(df_summary)))
    col_chart.set_categories(Reference(ws, min_col=1, min_row=col_row_start+2, max_row=col_row_start+1+len(df_summary)))
    ws.add_chart(col_chart, f"E{col_row_start+2}")

    run_cols = [col for col in df_summary.columns if col.startswith("run-")]
    if run_cols:
        line_row_start = col_row_start + len(df_summary) + 6
        ws.append([])
        ws.append(["Test Name"] + run_cols)
        for row in df_summary[["Test Name"] + run_cols].itertuples(index=False):
            ws.append(list(row))

        line_chart = LineChart()
        line_chart.title = "Durations Across Runs"
        line_chart.x_axis.title = "Test Name"
        line_chart.y_axis.title = "Duration (mins)"
        data_start = line_row_start + 2
        data_end = data_start + len(df_summary) - 1
        line_chart.add_data(Reference(ws, min_col=2, max_col=1+len(run_cols), min_row=data_start-1, max_row=data_end), titles_from_data=True)
        line_chart.set_categories(Reference(ws, min_col=1, min_row=data_start, max_row=data_end))
        ws.add_chart(line_chart, f"E{line_row_start+2}")

    wb.save(excel_path)

def extract_all_reports_from_folder(folder_path):
    log_path = os.path.join(folder_path, "processed_files.csv")
    output_excel = os.path.join(folder_path, "Test_Run_Details.xlsx")

    if REPROCESS_ALL:
        reset_processed_log(log_path)

    all_data = []
    processed = load_processed_files(log_path)

    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(".html") and file not in processed:
                html_path = os.path.join(root, file)
                test_info = extract_tests_from_html(html_path)
                if test_info:
                    all_data.extend(test_info)
                    save_processed_file(log_path, file)

    if all_data:
        df = pd.DataFrame(all_data)
        df["Start Time"] = pd.to_datetime(df["Start Time"], errors="coerce")
        df["End Time"] = pd.to_datetime(df["End Time"], errors="coerce")
        df["Duration (mins)"] = ((df["End Time"] - df["Start Time"]).dt.total_seconds() / 60).clip(lower=0).round(2)

        df_summary = summarize_test_durations(df)

        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="test run details")
            df_summary.to_excel(writer, index=False, sheet_name="summary report")

        add_charts_to_workbook(output_excel, df, df_summary)
        print(f"✅ Report generated at: {output_excel}")
    else:
        print("⚠️ No new HTML reports found.")

if __name__ == "__main__":
    folder_to_scan = "Reports"  # Change this to your folder
    extract_all_reports_from_folder(folder_to_scan)
