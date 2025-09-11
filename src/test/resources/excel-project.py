#!/usr/bin/env python3
"""
Excel Description Processor
==========================

A comprehensive solution for processing Excel files with pipe-separated description columns.
This script reads an Excel file, parses the Description column by splitting on the '|' symbol,
and populates the corresponding columns with structured test steps.

Features:
- Reads Excel files and identifies description columns
- Parses pipe-separated test steps
- Updates columns B & C with structured design steps and expected results
- Preserves existing data
- Provides detailed logging and verification
- Auto-activates virtual environment if needed

Author: AI Assistant
Date: 2025
"""

import sys
import os
import subprocess

def ensure_virtual_environment():
    """
    Ensure we're running in the virtual environment with required packages.
    If not, try to activate it and re-run the script.
    """
    # Check if we're already in a virtual environment
    if hasattr(sys, 'real_prefix') or (hasattr(sys, 'base_prefix') and sys.base_prefix != sys.prefix):
        # We're in a virtual environment, continue normally
        return True
    
    # Check if pandas is available (our main dependency)
    try:
        import pandas
        return True
    except ImportError:
        pass
    
    # Try to find and activate virtual environment
    script_dir = os.path.dirname(os.path.abspath(__file__))
    venv_path = os.path.join(script_dir, 'venv')
    
    if os.path.exists(venv_path):
        print("🔧 Virtual environment detected but not active. Activating and re-running...")
        
        # Determine the activation script path based on OS
        if os.name == 'nt':  # Windows
            activate_script = os.path.join(venv_path, 'Scripts', 'python.exe')
        else:  # Unix/Linux/macOS
            activate_script = os.path.join(venv_path, 'bin', 'python')
        
        if os.path.exists(activate_script):
            # Re-run the script with the virtual environment Python
            cmd = [activate_script] + sys.argv
            result = subprocess.run(cmd)
            sys.exit(result.returncode)
    
    # If we get here, we couldn't find or activate the virtual environment
    print("❌ Virtual environment not found or pandas not installed!")
    print("Please run the following commands first:")
    print("  python3 -m venv venv")
    print("  source venv/bin/activate  # On Windows: venv\\Scripts\\activate")
    print("  pip install -r requirements.txt")
    sys.exit(1)

# Ensure virtual environment is active before importing other packages
ensure_virtual_environment()

# Now we can safely import the required packages
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from typing import Tuple, Optional

# ========================================
# CONFIGURATION - UPDATE THESE VARIABLES
# ========================================
EXCEL_FILE = "Book1.xlsx"          # Excel file to process (change this to your file name)
SHEET_NAME = "Input"               # Sheet name containing the data (change if different)
CREATE_BACKUP = False              # Whether to create backup before processing (True/False)
FORCE_REPROCESS = True             # Process ALL rows even if columns B & C have data (True/False)
INDIVIDUAL_RECORDS = False         # Create individual rows for each step (True/False)

# Examples:
# EXCEL_FILE = "MyTestData.xlsx"
# SHEET_NAME = "TestCases" 
# CREATE_BACKUP = False
# FORCE_REPROCESS = False          # Only process empty cells (default behavior)
# ========================================

class ExcelDescriptionProcessor:
    """Main class for processing Excel description columns."""
    
    def __init__(self, file_path: str, sheet_name: str = 'Input', force_reprocess: bool = False, individual_records: bool = False):
        """
        Initialize the processor.
        
        Args:
            file_path (str): Path to the Excel file
            sheet_name (str): Name of the sheet to process (default: 'Input')
            force_reprocess (bool): Process all rows even if columns B & C have data (default: False)
            individual_records (bool): Create individual rows for each step (default: False)
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.force_reprocess = force_reprocess
        self.individual_records = individual_records
        self.backup_created = False
    
    def create_backup(self) -> bool:
        """
        Create a backup of the original Excel file.
        
        Returns:
            bool: True if backup was created successfully
        """
        try:
            backup_path = f"{os.path.splitext(self.file_path)[0]}_backup.xlsx"
            import shutil
            shutil.copy2(self.file_path, backup_path)
            print(f"✅ Backup created: {backup_path}")
            self.backup_created = True
            return True
        except Exception as e:
            print(f"⚠️  Warning: Could not create backup: {e}")
            return False
    
    def examine_excel_structure(self) -> bool:
        """
        Examine the Excel file structure and display information.
        
        Returns:
            bool: True if examination was successful
        """
        try:
            print("📋 Examining Excel File Structure")
            print("=" * 60)
            
            # Check available sheets
            xls = pd.ExcelFile(self.file_path)
            print(f"Available sheets: {xls.sheet_names}")
            
            if self.sheet_name not in xls.sheet_names:
                print(f"❌ Sheet '{self.sheet_name}' not found!")
                print(f"Available sheets: {xls.sheet_names}")
                return False
            
            # Read the target sheet
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            print(f"\nSheet: {self.sheet_name}")
            print(f"Shape: {df.shape}")
            print(f"Columns: {df.columns.tolist()}")
            
            # Check for required columns
            if 'Description' not in df.columns:
                print("❌ 'Description' column not found!")
                return False
            
            print("\n📊 Current Data Preview:")
            print("-" * 40)
            for idx, row in df.iterrows():
                desc_col_b = row.get('Description (Design Steps)', 'N/A')
                desc_col_c = row.get('Description (Expected Result)', 'N/A')
                
                print(f"Row {idx + 1}:")
                print(f"  Column A (Description): {'✓ Has data' if pd.notna(row['Description']) else '✗ Empty'}")
                print(f"  Column B (Design Steps): {'✓ Has data' if pd.notna(desc_col_b) else '✗ Empty (will be processed)'}")
                print(f"  Column C (Expected Result): {'✓ Has data' if pd.notna(desc_col_c) else '✗ Empty (will be processed)'}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error examining Excel structure: {e}")
            return False
    
    def parse_description_to_steps(self, description_text: str) -> Tuple[str, str]:
        """
        Parse the description text and extract structured steps as single cell values.
        Handles multi-line content within pipe-separated fields as raw text.
        
        Args:
            description_text (str): The pipe-separated description text
            
        Returns:
            tuple: (design_steps, expected_results) - single cell values, not combined
        """
        if pd.isna(description_text) or not description_text:
            return "", ""
        
        design_steps = []
        expected_results = []
        
        # Convert to string and normalize line endings
        text = str(description_text).replace('\r\n', '\n').replace('\r', '\n')
        
        # Split into potential step blocks by looking for step patterns
        # Use regex to find step patterns like "|Step-1|" or "| Step-1|"
        import re
        
        # Find all step patterns and their positions (exclude header patterns)
        step_pattern = r'\|\s*Step[-\d][^|]*\|'  # Match Step-1, Step1, etc. but not "Step No"
        step_matches = list(re.finditer(step_pattern, text, re.IGNORECASE))
        
        if not step_matches:
            # Fallback to line-by-line parsing if no step patterns found
            return self._parse_line_by_line(text)
        
        # Process each step block
        for i, match in enumerate(step_matches):
            # Get the start of this step
            start_pos = match.start()
            
            # Find the end position (start of next step or end of text)
            if i + 1 < len(step_matches):
                end_pos = step_matches[i + 1].start()
            else:
                end_pos = len(text)
            
            # Extract the complete step block
            step_block = text[start_pos:end_pos].strip()
            
            # Parse this step block
            parsed_step = self._parse_step_block(step_block)
            if parsed_step:
                design_steps.append(parsed_step['design'])
                expected_results.append(parsed_step['result'])
        
        # Return the results
        if design_steps:
            design_steps_text = '\n'.join(design_steps)
            expected_results_text = '\n'.join(expected_results)
        else:
            design_steps_text = ""
            expected_results_text = ""
        
        return design_steps_text, expected_results_text
    
    def _parse_step_block(self, step_block: str) -> dict:
        """
        Parse a single step block that may contain multiple pipes within content.
        Strategy: Everything from Step-X to next Step-Y is one block, split intelligently.
        
        Args:
            step_block (str): A block of text containing one step with potentially multiple pipes
            
        Returns:
            dict: Parsed step information or None if parsing fails
        """
        try:
            # Remove leading/trailing whitespace
            step_block = step_block.strip()
            
            # Remove leading and trailing pipes if they exist
            if step_block.startswith('|'):
                step_block = step_block[1:]
            if step_block.endswith('|'):
                step_block = step_block[:-1]
            
            # Split by pipe to get all parts
            all_parts = [part.strip() for part in step_block.split('|')]
            
            # Filter out empty parts
            parts = [part for part in all_parts if part]
            
            if len(parts) < 2:
                print(f"  ⚠️  Not enough parts in step block: {parts}")
                return None
            
            # First part is always the step number
            step_num = parts[0].strip()
            
            # Now we need to intelligently split the remaining parts into design and expected
            remaining_parts = parts[1:]
            
            # Strategy: Find the midpoint or use pattern recognition
            design_steps, expected_results = self._split_content_intelligently(remaining_parts)
            
            # Clean up whitespace while preserving structure
            design_text = ' '.join(design_steps.split()) if design_steps else ""
            expected_text = ' '.join(expected_results.split()) if expected_results else ""
            
            return {
                'design': f"{step_num}: {design_text}",
                'result': f"{step_num}: {expected_text}"
            }
            
        except Exception as e:
            print(f"  ⚠️  Warning: Could not parse step block: {step_block[:50]}... Error: {e}")
        
        return None
    
    def _split_content_intelligently(self, parts: list) -> tuple:
        """
        Intelligently split the content parts into design steps and expected results.
        
        Args:
            parts (list): List of content parts between pipes
            
        Returns:
            tuple: (design_steps, expected_results)
        """
        if not parts:
            return "", ""
        
        if len(parts) == 1:
            # Only one part, use it as design step
            return parts[0], ""
        
        # Look for patterns that might indicate expected results
        expected_keywords = ['should', 'expected', 'result', 'verify', 'confirm', 'display', 'show']
        
        # Find potential split points
        split_index = None
        
        # Strategy 1: Look for keywords that indicate expected results
        for i, part in enumerate(parts):
            part_lower = part.lower()
            if any(keyword in part_lower for keyword in expected_keywords):
                split_index = i
                break
        
        # Strategy 2: If no keywords found, split roughly in the middle
        if split_index is None:
            if len(parts) == 2:
                split_index = 1  # First part is design, second is expected
            else:
                split_index = len(parts) // 2  # Split in middle
        
        # Split the parts
        design_parts = parts[:split_index] if split_index > 0 else parts[:1]
        expected_parts = parts[split_index:] if split_index < len(parts) else []
        
        # Join the parts
        design_text = ' | '.join(design_parts) if design_parts else ""
        expected_text = ' | '.join(expected_parts) if expected_parts else ""
        
        return design_text, expected_text
    
    def _parse_line_by_line(self, text: str) -> Tuple[str, str]:
        """
        Fallback method to parse line by line (original logic).
        
        Args:
            text (str): The description text
            
        Returns:
            tuple: (design_steps, expected_results)
        """
        design_steps = []
        expected_results = []
        
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            # Skip header line and empty lines
            if not line or 'Step No' in line:
                continue
                
            # Split by pipe and extract step information
            if '|' in line:
                parts = [part.strip() for part in line.split('|') if part.strip()]
                
                # Expected format after splitting: ['Step-X', 'Action', 'Expected Result']
                if len(parts) >= 3:
                    step_num = parts[0]
                    action = parts[1]
                    result = parts[2]
                    
                    # Store with step numbers as before
                    design_steps.append(f"{step_num}: {action.strip()}")
                    expected_results.append(f"{step_num}: {result.strip()}")
        
        if design_steps:
            return '\n'.join(design_steps), '\n'.join(expected_results)
        else:
            return "", ""
    
    def parse_description_to_individual_steps(self, description_text: str) -> list:
        """
        Parse the description text and extract individual step records.
        
        Args:
            description_text (str): The pipe-separated description text
            
        Returns:
            list: List of dictionaries containing individual step information
        """
        if pd.isna(description_text) or not description_text:
            return []
        
        individual_steps = []
        
        # Split by newlines to get each step line
        lines = str(description_text).split('\n')
        
        for line in lines:
            line = line.strip()
            # Skip header line and empty lines
            if not line or 'Step No' in line:
                continue
                
            # Split by pipe and extract step information
            if '|' in line:
                parts = [part.strip() for part in line.split('|') if part.strip()]
                
                # Expected format after splitting: ['Step-X', 'Action', 'Expected Result']
                if len(parts) >= 3:
                    step_info = {
                        'step_number': parts[0],
                        'design_step': parts[1],
                        'expected_result': parts[2]
                    }
                    individual_steps.append(step_info)
        
        return individual_steps
    
    def update_excel_columns(self) -> bool:
        """
        Update the Excel file by populating columns B and C based on Description column.
        
        Returns:
            bool: True if update was successful
        """
        try:
            print("\n🔄 Processing Excel File")
            print("=" * 60)
            
            # Load the workbook using openpyxl for better control
            wb = load_workbook(self.file_path)
            ws = wb[self.sheet_name]
            
            # Read the current data using pandas for easier manipulation
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            
            updates_made = 0
            
            # Process each row
            for idx, row in df.iterrows():
                description = row['Description']
                current_design_steps = row.get('Description (Design Steps)')
                current_expected_result = row.get('Description (Expected Result)')
                
                # Determine if we should process this row
                should_process = False
                if self.force_reprocess:
                    # Force reprocess mode - process all rows
                    should_process = True
                    process_reason = "Force reprocess enabled"
                else:
                    # Normal mode - only process if columns B or C are empty
                    if pd.isna(current_design_steps) or pd.isna(current_expected_result):
                        should_process = True
                        process_reason = "Empty columns detected"
                
                if should_process:
                    print(f"\n📝 Processing Row {idx + 1}: ({process_reason})")
                    
                    # Parse the description
                    design_steps, expected_results = self.parse_description_to_steps(description)
                    
                    if design_steps or expected_results:
                        # Update the Excel cells (using 1-based indexing for openpyxl)
                        excel_row = idx + 2  # +2 because Excel is 1-indexed and we have headers
                        
                        # Update Column B if we have design steps
                        if design_steps:
                            if self.force_reprocess or pd.isna(current_design_steps):
                                ws.cell(row=excel_row, column=2, value=design_steps)  # Column B
                                action = "Updated" if self.force_reprocess and pd.notna(current_design_steps) else "Added"
                                print(f"  ✅ {action} Column B (Design Steps)")
                                print(f"     Content: {design_steps.replace(chr(10), ' | ')}")  # Show on one line
                        
                        # Update Column C if we have expected results
                        if expected_results:
                            if self.force_reprocess or pd.isna(current_expected_result):
                                ws.cell(row=excel_row, column=3, value=expected_results)  # Column C
                                action = "Updated" if self.force_reprocess and pd.notna(current_expected_result) else "Added"
                                print(f"  ✅ {action} Column C (Expected Results)")
                                print(f"     Content: {expected_results.replace(chr(10), ' | ')}")  # Show on one line
                        
                        updates_made += 1
                    else:
                        print(f"  ⚠️  No valid steps found in description")
                else:
                    print(f"Row {idx + 1}: ✓ Already has data in columns B & C (skipped)")
            
            if updates_made > 0:
                # Save the workbook
                wb.save(self.file_path)
                print(f"\n✅ Excel file updated successfully!")
                print(f"   File: {self.file_path}")
                print(f"   Rows updated: {updates_made}")
            else:
                print(f"\n📋 No updates needed - all rows already have data in columns B & C")
            
            return True
            
        except Exception as e:
            print(f"❌ Error updating Excel file: {e}")
            return False
    
    def create_individual_records(self) -> bool:
        """
        Create individual records for each step in a new sheet.
        
        Returns:
            bool: True if creation was successful
        """
        try:
            print("\n📋 Creating Individual Step Records")
            print("=" * 60)
            
            # Read the original data
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            
            # Create list to store individual records
            individual_records = []
            
            for idx, row in df.iterrows():
                description = row['Description']
                
                print(f"\n📝 Processing Row {idx + 1} for individual records:")
                
                # Parse individual steps
                steps = self.parse_description_to_individual_steps(description)
                
                if steps:
                    for step_idx, step in enumerate(steps):
                        record = {
                            'Original_Row': idx + 1,
                            'Step_Number': step['step_number'],
                            'Design_Step': step['design_step'],
                            'Expected_Result': step['expected_result'],
                            'Original_Description': description
                        }
                        individual_records.append(record)
                        print(f"  ✅ Step {step['step_number']}: {step['design_step']} → {step['expected_result']}")
                else:
                    print(f"  ⚠️  No valid steps found")
            
            if individual_records:
                # Create DataFrame for individual records
                records_df = pd.DataFrame(individual_records)
                
                print(f"\n📊 Created {len(individual_records)} individual step records")
                
                # Load workbook and add new sheet
                from openpyxl import load_workbook
                wb = load_workbook(self.file_path)
                
                # Create new sheet name
                new_sheet_name = f"{self.sheet_name}_Individual_Steps"
                
                # Remove sheet if it already exists
                if new_sheet_name in wb.sheetnames:
                    wb.remove(wb[new_sheet_name])
                    print(f"  ♻️  Replaced existing sheet: {new_sheet_name}")
                
                # Add the new sheet with individual records
                with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a') as writer:
                    records_df.to_excel(writer, sheet_name=new_sheet_name, index=False)
                
                print(f"✅ Individual records saved to sheet: {new_sheet_name}")
                print(f"   Total records: {len(individual_records)}")
                
                # Show preview of the records
                print(f"\n📋 Preview of Individual Records:")
                print("-" * 60)
                for i, record in enumerate(individual_records[:5]):  # Show first 5
                    print(f"Record {i+1}: {record['Step_Number']} | {record['Design_Step']} | {record['Expected_Result']}")
                
                if len(individual_records) > 5:
                    print(f"... and {len(individual_records) - 5} more records")
                
                return True
            else:
                print("❌ No individual records could be created")
                return False
                
        except Exception as e:
            print(f"❌ Error creating individual records: {e}")
            return False
    
    def verify_results(self) -> bool:
        """
        Verify the results after processing.
        
        Returns:
            bool: True if verification was successful
        """
        try:
            print("\n🔍 Verifying Results")
            print("=" * 60)
            
            # Read the updated Excel file
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            
            print(f"Final data shape: {df.shape}")
            print(f"Columns: {df.columns.tolist()}")
            
            print("\n📋 Final Results Summary:")
            print("-" * 40)
            
            for idx, row in df.iterrows():
                design_steps = row.get('Description (Design Steps)')
                expected_results = row.get('Description (Expected Result)')
                
                print(f"\nRow {idx + 1}:")
                if pd.notna(design_steps):
                    steps_preview = str(design_steps).replace('\n', ' | ')
                    print(f"  Column B: ✓ {steps_preview}")
                else:
                    print(f"  Column B: ✗ Empty")
                
                if pd.notna(expected_results):
                    results_preview = str(expected_results).replace('\n', ' | ')
                    print(f"  Column C: ✓ {results_preview}")
                else:
                    print(f"  Column C: ✗ Empty")
            
            return True
            
        except Exception as e:
            print(f"❌ Error verifying results: {e}")
            return False
    
    def process(self, create_backup: bool = True) -> bool:
        """
        Main processing method that runs the complete workflow.
        
        Args:
            create_backup (bool): Whether to create a backup before processing
            
        Returns:
            bool: True if processing was successful
        """
        print("🚀 Excel Description Processor")
        print("=" * 60)
        print(f"File: {self.file_path}")
        print(f"Sheet: {self.sheet_name}")
        mode = "Individual Records" if self.individual_records else "Combined Steps"
        print(f"Mode: {mode}")
        if self.force_reprocess:
            print(f"Force Reprocess: Enabled")
        print("=" * 60)
        
        # Check if file exists
        if not os.path.exists(self.file_path):
            print(f"❌ Error: File '{self.file_path}' not found!")
            return False
        
        # Create backup if requested
        if create_backup:
            self.create_backup()
        
        # Step 1: Examine structure
        if not self.examine_excel_structure():
            return False
        
        # Step 2: Update columns (if not individual records mode)
        if not self.individual_records:
            if not self.update_excel_columns():
                return False
        
        # Step 3: Create individual records (if enabled)
        if self.individual_records:
            if not self.create_individual_records():
                return False
        
        # Step 4: Verify results
        if not self.verify_results():
            return False
        
        print("\n🎉 Processing completed successfully!")
        print("=" * 60)
        
        return True


def main():
    """Main function to run the Excel Description Processor."""
    
    # Use configuration variables from top of file, but allow command line override
    if len(sys.argv) > 1:
        file_path = sys.argv[1]  # Command line override
    else:
        file_path = EXCEL_FILE   # Use configuration variable
    
    if len(sys.argv) > 2:
        sheet_name = sys.argv[2] # Command line override
    else:
        sheet_name = SHEET_NAME  # Use configuration variable
    
    # Create processor instance
    processor = ExcelDescriptionProcessor(file_path, sheet_name, force_reprocess=FORCE_REPROCESS, individual_records=INDIVIDUAL_RECORDS)
    
    # Run the processing
    success = processor.process(create_backup=CREATE_BACKUP)
    
    if success:
        print("\n✅ All operations completed successfully!")
        print(f"📁 Updated file: {file_path}")
        if processor.backup_created:
            backup_path = f"{os.path.splitext(file_path)[0]}_backup.xlsx"
            print(f"💾 Backup available: {backup_path}")
    else:
        print("\n❌ Processing failed!")
        return 1
    
    return 0


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
